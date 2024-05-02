import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.root.geometry("800x600")
        self.root.configure(bg="#0f4c75")  # Set background color

        # Fonts
        self.default_font = ("Arial", 12)
        self.title_font = ("Arial", 18, "bold")

        # Data
        self.expenses = self.load_expenses()

        # UI Elements
        self.create_widgets()
        self.create_graphs()

    def create_widgets(self):
        # Title
        title_label = tk.Label(self.root, text="Expense Tracker", font=self.title_font, bg="#0f4c75", fg="#bbe1fa")
        title_label.pack(pady=20)

        # Expense Entry Section
        entry_frame = tk.Frame(self.root, bg="#0f4c75")
        entry_frame.pack()

        # Expense Amount
        amount_label = tk.Label(entry_frame, text="Amount:", font=self.default_font, bg="#0f4c75", fg="#bbe1fa")
        amount_label.grid(row=0, column=0, padx=10, pady=10)
        self.amount_entry = tk.Entry(entry_frame, font=self.default_font)
        self.amount_entry.grid(row=0, column=1, padx=10, pady=10)

        # Category
        category_label = tk.Label(entry_frame, text="Category:", font=self.default_font, bg="#0f4c75", fg="#bbe1fa")
        category_label.grid(row=0, column=2, padx=10, pady=10)
        self.category_entry = tk.Entry(entry_frame, font=self.default_font)
        self.category_entry.grid(row=0, column=3, padx=10, pady=10)

        # Date
        date_label = tk.Label(entry_frame, text="Date:", font=self.default_font, bg="#0f4c75", fg="#bbe1fa")
        date_label.grid(row=0, column=4, padx=10, pady=10)
        self.date_entry = DateEntry(entry_frame, font=self.default_font, background='darkblue', foreground='white', borderwidth=2)
        self.date_entry.grid(row=0, column=5, padx=10, pady=10)

        # Buttons Frame
        buttons_frame = tk.Frame(self.root, bg="#0f4c75")
        buttons_frame.pack()

        add_button = tk.Button(buttons_frame, text="Add Expense", font=self.default_font, command=self.add_expense)
        add_button.grid(row=0, column=0, padx=10, pady=10)

        edit_button = tk.Button(buttons_frame, text="Edit Expense", font=self.default_font, command=self.edit_expense)
        edit_button.grid(row=0, column=1, padx=10, pady=10)

        delete_button = tk.Button(buttons_frame, text="Delete Expense", font=self.default_font, command=self.delete_expense)
        delete_button.grid(row=0, column=2, padx=10, pady=10)

        # Expense Listbox
        self.expense_listbox = tk.Listbox(self.root, width=70, font=self.default_font, bg="#bbe1fa", fg="#0f4c75")
        self.expense_listbox.pack(pady=20)

        for expense in self.expenses:
            self.expense_listbox.insert(tk.END, expense)

    def create_graphs(self):
        # Date vs Category Graph
        self.fig_date_category, self.ax_date_category = plt.subplots(figsize=(5, 5))
        self.ax_date_category.set_xlabel('Date', color="#bbe1fa")
        self.ax_date_category.set_ylabel('Category', color="#bbe1fa")
        self.canvas_date_category = FigureCanvasTkAgg(self.fig_date_category, master=self.root)
        self.canvas_date_category.draw()
        self.canvas_date_category.get_tk_widget().pack(fill=tk.BOTH, side=tk.LEFT, expand=True)

        # Total Expense vs Category Graph
        self.fig_total_expense_category, self.ax_total_expense_category = plt.subplots(figsize=(5, 5))
        self.ax_total_expense_category.set_xlabel('Category', color="#bbe1fa")
        self.ax_total_expense_category.set_ylabel('Total Expense', color="#bbe1fa")
        self.canvas_total_expense_category = FigureCanvasTkAgg(self.fig_total_expense_category, master=self.root)
        self.canvas_total_expense_category.draw()
        self.canvas_total_expense_category.get_tk_widget().pack(fill=tk.BOTH, side=tk.LEFT, expand=True)

    def add_expense(self):
        amount = self.amount_entry.get()
        category = self.category_entry.get()
        date = self.date_entry.get()

        if amount and category and date:
            expense_text = f"{amount} | {category} | {date}"
            self.expense_listbox.insert(tk.END, expense_text)
            self.expenses.append(expense_text)
            self.update_graphs()
            self.save_expenses()
        else:
            messagebox.showerror("Error", "Please fill in all fields.")

    def edit_expense(self):
        selected_index = self.expense_listbox.curselection()
        if not selected_index:
            messagebox.showerror("Error", "Please select an expense to edit.")
            return

        amount = self.amount_entry.get()
        category = self.category_entry.get()
        date = self.date_entry.get()

        if amount and category and date:
            expense_text = f"{amount} | {category} | {date}"
            self.expense_listbox.delete(selected_index)
            self.expense_listbox.insert(selected_index, expense_text)
            self.expenses[selected_index[0]] = expense_text
            self.update_graphs()
            self.save_expenses()
        else:
            messagebox.showerror("Error", "Please fill in all fields.")

    def delete_expense(self):
        selected_index = self.expense_listbox.curselection()
        if not selected_index:
            messagebox.showerror("Error", "Please select an expense to delete.")
            return

        self.expense_listbox.delete(selected_index)
        del self.expenses[selected_index[0]]
        self.update_graphs()
        self.save_expenses()

    def save_expenses(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Amount", "Category", "Date"])

        for expense in self.expenses:
            expense_data = expense.split('|')
            ws.append([expense_data[0].strip(), expense_data[1].strip(), expense_data[2].strip()])

        wb.save("expenses.xlsx")

    def load_expenses(self):
        try:
            wb = openpyxl.load_workbook("expenses.xlsx")
            ws = wb.active
            expenses = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                amount, category, date = row
                expenses.append(f"{amount} | {category} | {date}")
            return expenses
        except FileNotFoundError:
            return []

    def update_graphs(self):
        categories = []
        amounts = []
        dates = []
        for expense in self.expenses:
            amount, category, date = expense.split('|')
            categories.append(category.strip())
            amounts.append(float(amount.strip()))
            dates.append(date.strip())

        # Date vs Category
        self.ax_date_category.clear()
        self.ax_date_category.bar(dates, categories, color='#3282b8')  # Change color
        self.ax_date_category.set_xlabel('Date', color="#bbe1fa")
        self.ax_date_category.set_ylabel('Category', color="#bbe1fa")
        self.fig_date_category.tight_layout()
        self.canvas_date_category.draw()

        # Total Expense vs Category
        category_expenses = {}
        for expense in self.expenses:
            amount, category, _ = expense.split('|')
            category = category.strip()
            amount = float(amount.strip())
            category_expenses[category] = category_expenses.get(category, 0) + amount

        categories = list(category_expenses.keys())
        total_expenses = list(category_expenses.values())

        self.ax_total_expense_category.clear()
        self.ax_total_expense_category.bar(categories, total_expenses, color='#5e60ce')  # Change color
        self.ax_total_expense_category.set_xlabel('Category', color="#bbe1fa")
        self.ax_total_expense_category.set_ylabel('Total Expense', color="#bbe1fa")
        self.fig_total_expense_category.tight_layout()
        self.canvas_total_expense_category.draw()

def main():
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()

if __name__ == "__main__":
    main()
